# -*- coding: utf-8 -*-
"""P4 client: run p4 describe for each CL and parse changed files, submit time, and diff line info."""

import os
import re
import subprocess
import tempfile
import warnings
from typing import List, Optional, Tuple

# 用于解析 p4 describe 第一行的提交时间：Change 4875980 by user@client on 2026/03/04 21:20:36
_RE_CHANGE_ON = re.compile(r"Change\s+\d+\s+by\s+\S+\s+on\s+(.+)", re.I)
# 解析 diff 中的 @@ -old_start,old_count +new_start,new_count @@
_RE_DIFF_HUNK = re.compile(r"@@\s+-(\d+)(?:,(\d+))?\s+\+(\d+)(?:,(\d+))?\s+@@")
# 文本 diff 每个 hunk 最多展示的删/增行数（用于「修改内容」摘要）
_DIFF_SNIPPET_MAX_MINUS = 3
_DIFF_SNIPPET_MAX_PLUS = 3
_DIFF_LINE_MAX_LEN = 100
# 每个文件最多展示带内容的 hunk 数，超出只显示行号
_DIFF_HUNKS_WITH_SNIPPETS = 3

# xlsx/xlsm 单元格内容在摘要中显示的最大长度（避免单条过长；超出部分用 … 表示）
_EXCEL_CELL_DISPLAY_MAX = 120
# xlsx/xlsm 变更条数最多展示几条（超出则写「等共 N 处变更」）
_EXCEL_MAX_CHANGES_SHOW = 20

# 常见二进制/Office 扩展名：这些文件 P4 不会产生文本 diff，无法显示「第几行」变更
_BINARY_EXTENSIONS = frozenset(
    "xlsx xls doc docx ppt pptx pdf zip rar 7z png jpg jpeg gif bmp ico webp "
    "mp3 wav mp4 avi mov wmv dll so dylib exe bin".split()
)


def _no_diff_hint(depot_path: str, p4_marked_binary: bool = False) -> str:
    """对无行级 diff 的文件给出更易懂的说明（如 xlsx 等二进制）。"""
    ext = (depot_path.split(".")[-1].lower() if "." in depot_path else "").split("#")[0]
    if ext in _BINARY_EXTENSIONS:
        return "（二进制/Office 文件如 xlsx，无行级 diff，仅记录文件变更）"
    if p4_marked_binary:
        return "（P4 将该文件标为 binary，未输出文本 diff）"
    return "（未解析到行级 diff：可能 P4 标为 binary、或为空/集成变更、或编码异常）"


def _excel_change_summary(depot_path: str, new_rev: int, cwd: Optional[str] = None) -> Optional[str]:
    """
    对 xlsx/xlsm 用 p4 print 取新旧两版，用 openpyxl 比较单元格，返回可读的变更摘要。
    若未安装 openpyxl 或比较失败则返回 None。
    """
    try:
        import openpyxl
    except ImportError:
        return None
    if new_rev <= 1:
        return "（新文件，无旧版可对比）"
    old_rev = new_rev - 1
    spec_old = f"{depot_path}#{old_rev}"
    spec_new = f"{depot_path}#{new_rev}"
    cmd_old = ["p4", "print", "-q", spec_old]
    cmd_new = ["p4", "print", "-q", spec_new]
    try:
        r_old = subprocess.run(cmd_old, cwd=cwd, capture_output=True, timeout=30)
        r_new = subprocess.run(cmd_new, cwd=cwd, capture_output=True, timeout=30)
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return None
    if r_old.returncode != 0 or r_new.returncode != 0:
        return None
    data_old = r_old.stdout
    data_new = r_new.stdout
    if not data_old or not data_new:
        return None
    suffix = ".xlsx" if depot_path.lower().endswith(".xlsx") else ".xlsm"
    changes = []
    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as f_old:
            f_old.write(data_old)
            path_old = f_old.name
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as f_new:
            f_new.write(data_new)
            path_new = f_new.name
        try:
            # 抑制 openpyxl 对 WMF 等不支持图片格式的警告（我们只比较单元格数据，不依赖图片）
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                wb_old = openpyxl.load_workbook(path_old, data_only=True)
                wb_new = openpyxl.load_workbook(path_new, data_only=True)
            sheets_old = set(wb_old.sheetnames)
            sheets_new = set(wb_new.sheetnames)
            for name in sorted(sheets_new | sheets_old):
                if name not in sheets_old:
                    changes.append(f"【{name}】新增工作表")
                    continue
                if name not in sheets_new:
                    changes.append(f"【{name}】删除工作表")
                    continue
                ws_old = wb_old[name]
                ws_new = wb_new[name]
                cells_old = set()
                for row in ws_old.iter_rows():
                    for c in row:
                        cells_old.add((c.row, c.column))
                for row in ws_new.iter_rows():
                    for c in row:
                        coord = (c.row, c.column)
                        addr = openpyxl.utils.get_column_letter(c.column) + str(c.row)
                        new_val = c.value
                        if coord in cells_old:
                            old_val = ws_old.cell(c.row, c.column).value
                            if old_val != new_val:
                                max_len = _EXCEL_CELL_DISPLAY_MAX
                                old_s = "" if old_val is None else str(old_val)[:max_len]
                                new_s = "" if new_val is None else str(new_val)[:max_len]
                                if old_val is not None and len(str(old_val)) > max_len:
                                    old_s += "…"
                                if new_val is not None and len(str(new_val)) > max_len:
                                    new_s += "…"
                                changes.append(f"【{name}】{addr}: 「{old_s}」→「{new_s}」")
                        else:
                            max_len = _EXCEL_CELL_DISPLAY_MAX
                            new_s = "" if new_val is None else str(new_val)[:max_len]
                            if new_val is not None and len(str(new_val)) > max_len:
                                new_s += "…"
                            changes.append(f"【{name}】{addr}: 新增「{new_s}」")
        finally:
            try:
                os.unlink(path_old)
            except OSError:
                pass
            try:
                os.unlink(path_new)
            except OSError:
                pass
    except Exception:
        return None
    if not changes:
        return "（内容未变或无法解析）"
    parts = changes[:_EXCEL_MAX_CHANGES_SHOW]
    summary = "；".join(parts)
    if len(changes) > _EXCEL_MAX_CHANGES_SHOW:
        summary += f" 等共 {len(changes)} 处变更"
    return summary


def describe_cl(cl: int, cwd: Optional[str] = None) -> tuple[List[str], Optional[str]]:
    """
    Run `p4 describe -s <CL>` and return (list of depot paths, error_message).

    -s: short output, one line per file.
    """
    cmd = ["p4", "describe", "-s", str(cl)]
    try:
        result = subprocess.run(
            cmd,
            cwd=cwd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=60,
        )
    except FileNotFoundError:
        return [], "未找到 p4 命令，请确保 Perforce 已在 PATH 中"
    except subprocess.TimeoutExpired:
        return [], f"p4 describe 超时: CL {cl}"

    out = result.stdout or ""
    err = result.stderr or ""
    if result.returncode != 0:
        return [], f"CL {cl}: p4 describe 失败 (returncode={result.returncode})\n{err[:300]}"

    # 解析 depot 路径。p4 describe -s 中文件行格式为: "... //FantasyWorld/.../file.cs#59 edit"
    # 不要跳过以 "..." 开头的行（那里才有真实路径）；只跳过纯 "..." 或 "Change ..." 等
    paths = []
    for line in out.splitlines():
        line = line.strip()
        if not line or line.startswith("Change ") or line == "...":
            continue
        # 提取该行中所有 // 开头的 depot 路径（到空格或 # 为止）
        for m in re.finditer(r"//[^\s#]+", line):
            p = m.group(0)
            if len(p) <= 2:
                continue
            # 排除描述里的 URL（如 [Jira] https://xindong.atlassian.net/... 被误匹配为 //xindong.atlassian.net/...）
            if ".atlassian.net" in p or "http" in p.lower():
                continue
            paths.append(p)
    return paths, None


def _parse_describe_full(out: str, cl: int) -> tuple[str, List[str], List[Tuple[str, str]], dict]:
    """
    解析 p4 describe（完整输出）：提交时间、文件路径列表、每个文件的变更行摘要、路径对应的 rev。
    Returns: (submit_time, paths, [(path, line_summary), ...], path_to_rev)
    """
    path_to_rev = {}  # path -> revision number (from #59 in "... //path/file.xlsx#59 edit")
    submit_time = ""
    paths = []
    file_line_summaries = []  # [(path, summary), ...]

    lines = out.splitlines()
    # 第一行：Change 4875980 by user@client on 2026/03/04 21:20:36
    for line in lines[:5]:
        m = _RE_CHANGE_ON.match(line.strip())
        if m:
            submit_time = m.group(1).strip()
            break

    current_path = None
    # list of (new_start, new_count, old_start, old_count, [(sign, text), ...])
    current_diffs: List[Tuple[int, int, int, int, List[Tuple[str, str]]]] = []
    current_section_has_binary = False  # P4 在该文件段落里输出了 (binary) 等

    def _truncate_line(s: str) -> str:
        s = (s or "").replace("\r", "").replace("\n", " ")
        return s[: _DIFF_LINE_MAX_LEN] + ("…" if len(s) > _DIFF_LINE_MAX_LEN else "")

    def flush_file():
        if not current_path:
            return
        if current_diffs:
            parts = []
            for idx, hunk in enumerate(current_diffs):
                ns, nc, os, oc = hunk[0], hunk[1], hunk[2], hunk[3]
                snippets = hunk[4] if len(hunk) > 4 else []
                line_info = f"第{ns}-{ns + nc - 1}行(+{nc}/-{oc})"
                if idx < _DIFF_HUNKS_WITH_SNIPPETS and snippets:
                    snip_strs = []
                    for sign, text in snippets:
                        t = _truncate_line(text)
                        snip_strs.append(f"{sign}{t}")
                    parts.append(line_info + ": " + " ".join(snip_strs))
                else:
                    parts.append(line_info)
            file_line_summaries.append(
                (current_path, "；".join(parts[:5]) + (" 等" if len(parts) > 5 else ""))
            )
        else:
            file_line_summaries.append((current_path, _no_diff_hint(current_path, current_section_has_binary)))

    def _is_new_file_line(ln: str) -> bool:
        s = ln.strip()
        if "..." not in s[:10]:
            return False
        return bool(re.search(r"//[^\s#]+", s) and ".atlassian.net" not in s and "http" not in s.lower())

    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        # 文件行: "... //depot/.../file#rev action"
        path_match = re.search(r"//[^\s#]+", stripped)
        if path_match and "..." in stripped[:10] and ".atlassian.net" not in stripped and "http" not in stripped.lower():
            p = path_match.group(0)
            if len(p) > 2:
                flush_file()
                paths.append(p)
                rev_m = re.search(r"#(\d+)", stripped)
                if rev_m:
                    path_to_rev[p] = int(rev_m.group(1))
                current_path = p
                current_diffs = []
                current_section_has_binary = False
        # P4 对 binary 文件会输出 (binary) 或 binary file，不输出 @@ diff
        if current_path and "binary" in stripped.lower():
            current_section_has_binary = True
        # diff 块 @@ -a,b +c,d @@，并收集后续 -/+ 行作为修改内容
        mm = _RE_DIFF_HUNK.search(stripped)
        if mm and current_path:
            old_start = int(mm.group(1))
            old_count = int(mm.group(2) or 1)
            new_start = int(mm.group(3))
            new_count = int(mm.group(4) or 1)
            snippets: List[Tuple[str, str]] = []
            j = i + 1
            n_minus, n_plus = 0, 0
            while j < len(lines):
                next_ln = lines[j]
                next_s = next_ln.strip()
                if _RE_DIFF_HUNK.search(next_s) or _is_new_file_line(next_ln):
                    break
                if next_ln.startswith("-") and not next_ln.startswith("---") and n_minus < _DIFF_SNIPPET_MAX_MINUS:
                    snippets.append(("-", next_ln[1:].strip()))
                    n_minus += 1
                elif next_ln.startswith("+") and not next_ln.startswith("+++") and n_plus < _DIFF_SNIPPET_MAX_PLUS:
                    snippets.append(("+", next_ln[1:].strip()))
                    n_plus += 1
                j += 1
            current_diffs.append((new_start, new_count, old_start, old_count, snippets))
            i = j - 1  # 已读取到 j，下一轮 i+1 会从 j 开始（下一段 @@ 或 path）
        i += 1
    flush_file()

    for p in paths:
        if not any(fp == p for fp, _ in file_line_summaries):
            file_line_summaries.append((p, _no_diff_hint(p, False)))
    return submit_time, paths, file_line_summaries, path_to_rev


def describe_cl_full(cl: int, cwd: Optional[str] = None) -> tuple[str, List[str], List[Tuple[str, str]], Optional[str]]:
    """
    执行 p4 describe <CL>（完整，非 -s），解析提交时间、文件列表、每个文件的变更行摘要。
    Returns: (submit_time, paths, [(path, line_summary)], error_message)
    """
    cmd = ["p4", "describe", str(cl)]
    try:
        result = subprocess.run(
            cmd,
            cwd=cwd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
        )
    except FileNotFoundError:
        return "", [], [], "未找到 p4 命令"
    except subprocess.TimeoutExpired:
        return "", [], [], f"p4 describe 超时: CL {cl}"

    out = result.stdout or ""
    err = result.stderr or ""
    if result.returncode != 0:
        return "", [], [], f"CL {cl}: p4 describe 失败\n{err[:200]}"

    try:
        submit_time, paths, file_line_summaries, path_to_rev = _parse_describe_full(out, cl)
    except Exception:
        # 解析失败时退化为只取路径（用 -s 结果）
        paths_short, _ = describe_cl(cl, cwd=cwd)
        return "", paths_short, [(p, "（解析失败）") for p in paths_short], None
    # 对 xlsx/xlsm 尝试用 openpyxl 生成单元格级变更摘要
    excel_ext = frozenset(("xlsx", "xlsm"))
    result_summaries = []
    for path, summary in file_line_summaries:
        ext = (path.split(".")[-1].lower() if "." in path else "").split("#")[0]
        if ext in excel_ext and summary == _no_diff_hint(path):
            rev = path_to_rev.get(path)
            if rev:
                excel_sum = _excel_change_summary(path, rev, cwd=cwd)
                if excel_sum:
                    summary = excel_sum
        result_summaries.append((path, summary))
    return submit_time, paths, result_summaries, None


# 方法2：用 p4 print 拉取变更文件内容，供 AI 分析（不依赖 agent 工作区）
_P4_PRINT_MAX_FILES = 15
_P4_PRINT_MAX_CHARS_PER_FILE = 2000
_P4_PRINT_MAX_TOTAL_CHARS = 8000


def get_file_contents_for_paths(
    depot_paths: List[str],
    cwd: Optional[str] = None,
    max_files: int = _P4_PRINT_MAX_FILES,
    max_chars_per_file: int = _P4_PRINT_MAX_CHARS_PER_FILE,
    max_total_chars: int = _P4_PRINT_MAX_TOTAL_CHARS,
) -> str:
    """
    对变更的 depot 路径用 p4 print 取文件内容，拼成一段文本供 AI 使用。
    跳过二进制/Office 等不可读类型；单文件与总长度截断，避免 prompt 过大。

    Returns:
        拼接后的内容，格式：--- depot_path ---\\n 内容\\n\\n；无内容或失败返回空字符串。
    """
    if not depot_paths:
        return ""
    lines = []
    total = 0
    for path in depot_paths[:max_files]:
        path = (path or "").strip().split("#")[0]
        if not path or not path.startswith("//"):
            continue
        ext = (path.split(".")[-1].lower() if "." in path else "").split("#")[0]
        
        # 如果是 xlsx/xlsm，我们不拉取二进制正文，而是尝试调用 _excel_change_summary 获取单元格变更摘要
        if ext in ("xlsx", "xlsm"):
            # 需要知道当前的 rev，这里我们假设它在 p4 files 或 p4 fstat 中能查到，或者简单地传 0 并不完全准确
            # 为了简单起见，我们直接用 p4 describe 里已经解析好的摘要（如果能传进来的话）
            # 但这里我们没有 rev，所以我们用 p4 fstat 查一下最新的 rev
            try:
                fstat_r = subprocess.run(["p4", "fstat", "-T", "headRev", path], cwd=cwd, capture_output=True, text=True, timeout=10)
                if fstat_r.returncode == 0 and fstat_r.stdout:
                    import re
                    rev_m = re.search(r'\.\.\.\s+headRev\s+(\d+)', fstat_r.stdout)
                    if rev_m:
                        rev = int(rev_m.group(1))
                        excel_summary = _excel_change_summary(path, rev, cwd=cwd)
                        if excel_summary:
                            block = f"--- {path} (Excel 变更摘要) ---\n{excel_summary}\n\n"
                            lines.append(block)
                            total += len(block)
            except Exception:
                pass
            continue
            
        if ext in _BINARY_EXTENSIONS:
            continue
        try:
            r = subprocess.run(
                ["p4", "print", "-q", path],
                cwd=cwd,
                capture_output=True,
                timeout=30,
            )
        except (FileNotFoundError, subprocess.TimeoutExpired):
            continue
        if r.returncode != 0:
            continue
        raw = r.stdout
        if not raw:
            continue
        try:
            text = raw.decode("utf-8", errors="replace")
        except Exception:
            continue
        if "\x00" in text or any(ord(c) < 32 and c not in "\n\r\t" for c in text[:100]):
            continue
        text = text.strip()
        if len(text) > max_chars_per_file:
            text = text[:max_chars_per_file] + "\n...（已截断）"
        block = f"--- {path} ---\n{text}\n\n"
        if total + len(block) > max_total_chars:
            remain = max_total_chars - total - 60
            if remain > 0:
                lines.append(f"--- {path} ---\n{text[:remain]}\n...（已截断）\n\n")
            break
        lines.append(block)
        total += len(block)
    return "".join(lines).strip() if lines else ""


def get_changed_files_for_cls(
    cl_list: List[int],
    cwd: Optional[str] = None,
) -> Tuple[List[Tuple[int, List[str], str, List[Tuple[str, str]]]], List[int], List[str]]:
    """
    For each CL, run p4 describe (full) and collect paths, submit time, and per-file line change summary.

    Returns:
        (files_by_cl, failed_cls, error_messages)
        - files_by_cl: list of (cl, paths, submit_time, file_line_summaries)
          file_line_summaries: list of (depot_path, line_summary_str)
        - failed_cls: list of CL numbers that failed
        - error_messages: list of error strings for failed CLs
    """
    files_by_cl = []
    failed_cls = []
    errors = []

    for cl in cl_list:
        submit_time, paths, file_line_summaries, err = describe_cl_full(cl, cwd=cwd)
        if err:
            failed_cls.append(cl)
            errors.append(err)
        else:
            files_by_cl.append((cl, paths, submit_time, file_line_summaries))

    return files_by_cl, failed_cls, errors


def _common_depot_prefix(paths: List[str]) -> str:
    """计算 depot 路径列表的最长公共前缀（按路径段）。"""
    if not paths:
        return ""
    cleaned = [p.strip() for p in paths if (p or "").strip()]
    if not cleaned:
        return ""
    segments_list = [p.split("/") for p in cleaned]
    common = []
    for i in range(min(len(s) for s in segments_list)):
        seg = segments_list[0][i]
        if all(s[i] == seg for s in segments_list):
            common.append(seg)
        else:
            break
    return "/".join(common) if common else ""


# 项目上下文：列举的目录/文件数量上限
_PROJECT_CONTEXT_MAX_DIRS = 80
_PROJECT_CONTEXT_MAX_FILES = 250


def get_project_context(depot_paths: List[str], cwd: Optional[str] = None) -> str:
    """
    根据变更涉及的 depot 路径，查询该区域的项目结构（子目录 + 文件示例），
    供 AI 结合「项目全貌」给出更全面的测试范围建议。

    Returns:
        一段可读的「项目结构」摘要；查询失败或路径为空则返回空字符串。
    """
    if not depot_paths:
        return ""
    prefix = _common_depot_prefix(depot_paths)
    if not prefix or prefix.count("/") < 2:
        return ""
    # 避免范围过大：至少取到变更路径的父级（如 //depot/Proj/Module）
    parts = prefix.split("/")
    if len(parts) <= 2:
        return ""

    lines = [f"项目 depot 区域（与本次变更相关）：{prefix}", ""]
    # 子目录：p4 dirs prefix/*
    spec_dirs = f"{prefix}/*"
    try:
        r = subprocess.run(
            ["p4", "dirs", spec_dirs],
            cwd=cwd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=15,
        )
        if r.returncode == 0 and (r.stdout or "").strip():
            dirs = [ln.strip() for ln in (r.stdout or "").splitlines() if ln.strip()][:_PROJECT_CONTEXT_MAX_DIRS]
            if dirs:
                lines.append("直接子目录：")
                for d in dirs:
                    short = d.split("/")[-1] if "/" in d else d
                    lines.append(f"  - {short}")
                lines.append("")
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass
    # 文件示例：p4 files prefix/...
    spec_files = f"{prefix}/..."
    try:
        r = subprocess.run(
            ["p4", "files", "-m", str(_PROJECT_CONTEXT_MAX_FILES), spec_files],
            cwd=cwd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=20,
        )
        if r.returncode == 0 and (r.stdout or "").strip():
            file_lines = [ln.strip() for ln in (r.stdout or "").splitlines() if ln.strip()]
            if file_lines:
                lines.append("该区域文件示例（用于了解模块与类型）：")
                for ln in file_lines[: _PROJECT_CONTEXT_MAX_FILES]:
                    # p4 files 输出形如：//depot/path#rev - action
                    path = ln.split("#")[0].strip().split("-")[0].strip()
                    if path:
                        lines.append(f"  {path}")
                if len(file_lines) > _PROJECT_CONTEXT_MAX_FILES:
                    lines.append(f"  ... 共约 {len(file_lines)} 个文件（已截断）")
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass
    return "\n".join(lines).strip() if len(lines) > 1 else ""


def get_game_project_structure(
    project_path: str,
    focus_subdirs: Optional[List[str]] = None,
    max_depth_root: int = 2,
    max_depth_focus: int = 3,
    max_chars: int = 2000,
) -> str:
    """
    从本机游戏工程目录收集目录结构摘要，供 AI 结合工程全貌给出游戏向测试范围。
    - 若 focus_subdirs 为空：只扫根目录下 max_depth_root 层（默认 2 层）。
    - 若 focus_subdirs 非空（通常由 P4 变更路径得出）：根目录扫 1 层；对每个 focus 子目录扫 max_depth_focus 层（默认 4 层），便于结合变更做更深结构。
    若路径不存在或无权限则返回空字符串。
    """
    try:
        from pathlib import Path
    except ImportError:
        return ""
    if not (project_path or "").strip():
        return ""
    root = Path(project_path.strip()).resolve()
    if not root.is_dir():
        return ""
    lines = [f"游戏工程本地路径：{root}", ""]
    try:
        # 根目录：扫 max_depth_root 层（或 focus 时只扫 1 层）
        depth_root = 1 if focus_subdirs else max_depth_root
        dirs_top = []
        dirs_second = []
        file_exts_root = {}
        for entry in sorted(root.iterdir())[:50]:
            if entry.name.startswith("."):
                continue
            if entry.is_dir():
                dirs_top.append(entry.name)
                if depth_root >= 2:
                    for sub in sorted(entry.iterdir())[:20]:
                        if sub.name.startswith("."):
                            continue
                        if sub.is_dir():
                            dirs_second.append(f"  {entry.name}/{sub.name}")
                        else:
                            ext = sub.suffix.lower() or "(无后缀)"
                            file_exts_root[ext] = file_exts_root.get(ext, 0) + 1
            else:
                ext = entry.suffix.lower() or "(无后缀)"
                file_exts_root[ext] = file_exts_root.get(ext, 0) + 1
        if dirs_top:
            lines.append("根目录一级：")
            lines.extend(f"  - {d}" for d in dirs_top[:25])
            lines.append("")
        if dirs_second:
            lines.append("根目录二级示例：")
            lines.extend(dirs_second[:20])
            lines.append("")
        if file_exts_root:
            summary = "、".join(f"{k}({c})" for k, c in sorted(file_exts_root.items(), key=lambda x: -x[1])[:15])
            lines.append(f"根目录文件类型示例：{summary}")
            lines.append("")

        # 根据 P4 变更聚焦的子目录：扫更深
        if focus_subdirs:
            seen = set()
            for rel in focus_subdirs[:8]:  # 最多 8 个 focus 目录，控制总长度
                rel = (rel or "").strip().replace("\\", "/").strip("/")
                if not rel or rel in seen:
                    continue
                seen.add(rel)
                focus_dir = root / rel
                if not focus_dir.is_dir():
                    continue
                lines.append(f"【P4 变更相关】{rel}/ （下 {max_depth_focus} 层）")
                _append_dir_tree(focus_dir, lines, max_depth=max_depth_focus, indent="  ", max_entries_per_dir=12)
                lines.append("")
        out = "\n".join(lines).strip()
        return out[:max_chars] if out else ""
    except (PermissionError, OSError):
        return ""


def _append_dir_tree(
    directory: "Path",
    lines: List[str],
    max_depth: int,
    indent: str,
    max_entries_per_dir: int,
    current_depth: int = 0,
) -> None:
    """递归列出目录结构，最多 max_depth 层。"""
    try:
        from pathlib import Path
    except ImportError:
        return
    if current_depth >= max_depth:
        return
    entries = sorted(directory.iterdir())[:max_entries_per_dir]
    for entry in entries:
        if entry.name.startswith("."):
            continue
        if entry.is_dir():
            lines.append(f"{indent}- {entry.name}/")
            _append_dir_tree(
                entry,
                lines,
                max_depth=max_depth,
                indent=indent + "  ",
                max_entries_per_dir=max_entries_per_dir,
                current_depth=current_depth + 1,
            )
        else:
            ext = entry.suffix.lower() or "(无后缀)"
            lines.append(f"{indent}- {entry.name} ({ext})")
    if len(list(directory.iterdir())) > max_entries_per_dir:
        lines.append(f"{indent}  ... 已截断")


def get_related_code_context(
    changed_paths: List[str],
    depot_prefix: str,
    cwd: Optional[str] = None,
    max_keywords: int = 3,
    max_related_files: int = 3,
    max_chars_per_file: int = 1500
) -> str:
    """
    通过提取变更文件中的关键字，使用 p4 grep 查找关联的代码文件，并返回其内容。
    供 AI 分析更准确的测试范围。
    """
    if not changed_paths or not depot_prefix:
        return ""

    # 1. 提取关键字：针对代码文件和配置表，提取文件名作为关键字
    keywords = []
    # 增加 xlsx/xlsm 支持，因为配置表的文件名通常也是代码中引用的类名/表名
    code_exts = {"cs", "lua", "ts", "js", "py", "cpp", "h", "xlsx", "xlsm"}
    for path in changed_paths:
        ext = (path.split(".")[-1].lower() if "." in path else "").split("#")[0]
        if ext in code_exts:
            # 提取文件名，例如 //depot/Client/CropComponent.cs -> CropComponent
            # 例如 //depot/Config/logic/TableCrop.xlsx -> TableCrop
            filename = path.split("/")[-1].split(".")[0]
            # 过滤掉太短或太常见的词，防止搜索爆炸
            if len(filename) > 3 and filename.lower() not in ["main", "manager", "config", "utils", "data", "table"]:
                keywords.append(filename)
                
    # 去重并限制关键字数量
    keywords = list(dict.fromkeys(keywords))[:max_keywords]
    
    if os.environ.get("FEISHU_REPORT_DOC_DEBUG"):
        print(f"[P4 Grep Debug] 提取到的搜索关键字: {keywords}", flush=True)

    if not keywords:
        return ""

    related_paths = set()
    
    # 2. 转变逻辑：放弃使用极度消耗性能的 p4 grep（搜索文件内容）
    # 改为使用 p4 files 搜索【文件名包含关键字】的代码文件。这种方式速度极快，且不会触发服务器的 revision limit 限制。
    search_scope_base = f"{depot_prefix}/Client" if "Client" not in depot_prefix else depot_prefix
    
    for kw in keywords:
        if len(related_paths) >= max_related_files:
            break
            
        # 清理关键字
        import re
        clean_kw = re.sub(r'[^a-zA-Z0-9_]', '', kw)
        if len(clean_kw) < 3:
            continue
            
        # P4 路径匹配区分大小写，我们尝试原词、首字母大写、全小写三种组合
        kw_variants = list(set([clean_kw, clean_kw.capitalize(), clean_kw.lower()]))
        
        for variant in kw_variants:
            if len(related_paths) >= max_related_files:
                break
                
            # 构造通配符路径，例如 //FantasyWorld/stage/Client/...*Crop*.cs
            spec = f"{search_scope_base}/...*{variant}*.cs"
            cmd = ["p4", "files", "-m", "5", spec]
            
            if os.environ.get("FEISHU_REPORT_DOC_DEBUG"):
                print(f"[P4 Files Debug] 正在按文件名搜索: {' '.join(cmd)}", flush=True)
                
            try:
                r = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True, timeout=10)
                if r.returncode == 0 and r.stdout:
                    for line in r.stdout.splitlines():
                        # p4 files 输出格式: //depot/path/file.cs#1 - add change 123 (text)
                        p = line.split("#")[0].strip()
                        if p and p.endswith(".cs") and p not in changed_paths:
                            related_paths.add(p)
                            if len(related_paths) >= max_related_files:
                                break
            except Exception as e:
                if os.environ.get("FEISHU_REPORT_DOC_DEBUG"):
                    print(f"[P4 Files Debug] 搜索异常: {e}", flush=True)
                continue

    if os.environ.get("FEISHU_REPORT_DOC_DEBUG"):
        print(f"[P4 Grep Debug] 找到的关联文件: {related_paths}", flush=True)

    if not related_paths:
        return ""

    # 3. 读取这些关联文件的内容
    lines = ["【以下是工程中引用了本次变更模块的关联代码文件，供推断影响范围参考】\n"]
    for path in list(related_paths)[:max_related_files]:
        if os.environ.get("FEISHU_REPORT_DOC_DEBUG"):
            print(f"[P4 Print Debug] 正在拉取关联文件内容: {path}", flush=True)
        try:
            r = subprocess.run(["p4", "print", "-q", path], cwd=cwd, capture_output=True, timeout=10)
            if r.returncode == 0 and r.stdout:
                text = r.stdout.decode("utf-8", errors="replace").strip()
                if len(text) > max_chars_per_file:
                    text = text[:max_chars_per_file] + "\n...（已截断）"
                lines.append(f"--- 关联文件: {path} ---\n{text}\n")
            elif r.returncode != 0 and os.environ.get("FEISHU_REPORT_DOC_DEBUG"):
                print(f"[P4 Print Debug] 拉取失败，returncode: {r.returncode}, stderr: {r.stderr}", flush=True)
        except Exception as e:
            if os.environ.get("FEISHU_REPORT_DOC_DEBUG"):
                print(f"[P4 Print Debug] 拉取异常: {e}", flush=True)
            continue

    return "\n".join(lines)

